import os
import pymysql
from pymysql import Error
from dotenv import load_dotenv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Load environment variables
load_dotenv()


# ============================================================================
# CONFIGURATION
# ============================================================================
DB_CONFIGS = {
    "host": os.getenv("MYSQL_GRDATA_HOST"),
    "port": int(os.getenv("MYSQL_DB_PORT", "3306")),
    "username": os.getenv("MYSQL_GRDATA_USER"),
    "password": os.getenv("MYSQL_GRDATA_PASSWORD", ""),
    "database": os.getenv("MYSQL_GRDATA_DB"),
    "table_name": None,
}


def get_mysql_db_client():
    """Return database configuration"""
    return DB_CONFIGS


def establish_connection_mysql_db():
    """
    Establish and return MySQL database connection with SSL fallback.
    Tries SSL connection first (like MySQL Workbench), falls back to non-SSL if needed.
    
    Returns:
        pymysql.connections.Connection: Active database connection or None
    """
    try:
        # First try with SSL (like MySQL Workbench does)
        print("Attempting connection with SSL...")
        try:
            connection = pymysql.connect(
                host=DB_CONFIGS["host"],
                port=DB_CONFIGS["port"],
                user=DB_CONFIGS["username"],
                password=DB_CONFIGS["password"],
                database=DB_CONFIGS["database"],
                charset='utf8mb4',
                cursorclass=pymysql.cursors.DictCursor,
                ssl={'ssl_mode': 'REQUIRED'}
            )
            print("✓ Connected with SSL")
        except Exception as ssl_error:
            print(f"SSL connection failed: {ssl_error}")
            print("\nAttempting connection without SSL...")
            # Try without SSL
            connection = pymysql.connect(
                host=DB_CONFIGS["host"],
                port=DB_CONFIGS["port"],
                user=DB_CONFIGS["username"],
                password=DB_CONFIGS["password"],
                database=DB_CONFIGS["database"],
                charset='utf8mb4',
                cursorclass=pymysql.cursors.DictCursor
            )
            print("✓ Connected without SSL")
        
        if connection.open:
            print(f"✓ Successfully connected to database: {DB_CONFIGS['database']}")
            return connection
            
    except Error as e:
        print(f"✗ Error connecting to MySQL database: {e}")
        print("\nTroubleshooting tips:")
        print("1. Verify your IP is whitelisted in the database server")
        print("2. Confirm the host IP is correct and accessible")
        print("3. Check username, password, and database name")
        print("4. Compare with MySQL Workbench connection settings")
        return None


def get_all_tables(connection):
    """Get list of all tables in the database"""
    try:
        with connection.cursor() as cursor:
            cursor.execute("SHOW TABLES")
            tables = [list(row.values())[0] for row in cursor.fetchall()]
        return tables
    except Error as e:
        print(f"✗ Error fetching tables: {e}")
        return []


def get_table_row_count(connection, table_name):
    """Get row count for a specific table"""
    try:
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT COUNT(*) as count FROM `{table_name}`")
            result = cursor.fetchone()
            return result['count']
    except Error as e:
        print(f"✗ Error getting row count for {table_name}: {e}")
        return 0


def get_table_columns(connection, table_name):
    """Get all column names for a table"""
    try:
        with connection.cursor() as cursor:
            cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
            columns = [row['Field'] for row in cursor.fetchall()]
        return columns
    except Error as e:
        print(f"✗ Error getting columns for {table_name}: {e}")
        return []


def find_timestamp_column(columns, patterns):
    """Find first matching timestamp column from list of patterns"""
    columns_lower = [col.lower() for col in columns]
    for pattern in patterns:
        if pattern.lower() in columns_lower:
            idx = columns_lower.index(pattern.lower())
            return columns[idx]
    return None


def get_max_timestamp(connection, table_name, column_name):
    """Get maximum value from a timestamp column"""
    if not column_name:
        return None
    
    try:
        with connection.cursor() as cursor:
            query = f"SELECT MAX(`{column_name}`) as max_ts FROM `{table_name}`"
            cursor.execute(query)
            result = cursor.fetchone()
            return result['max_ts'] if result['max_ts'] else None
    except Error as e:
        print(f"✗ Error getting max timestamp for {table_name}.{column_name}: {e}")
        return None


def collect_table_summaries(connection):
    """Collect summary information for all tables"""
    tables = get_all_tables(connection)
    summaries = []
    
    updated_patterns = ['updated_at', 'updatedAt', 'modified_at', 'modifiedAt', 'last_updated']
    created_patterns = ['created_at', 'createdAt', 'created', 'date_created']
    
    print(f"\nProcessing {len(tables)} tables...")
    
    for table_name in tables:
        print(f"  - Processing: {table_name}")
        
        # Get columns
        columns = get_table_columns(connection, table_name)
        
        # Find timestamp columns
        updated_col = find_timestamp_column(columns, updated_patterns)
        created_col = find_timestamp_column(columns, created_patterns)
        
        # Get values
        row_count = get_table_row_count(connection, table_name)
        max_updated = get_max_timestamp(connection, table_name, updated_col)
        max_created = get_max_timestamp(connection, table_name, created_col)
        
        summaries.append({
            'table_name': table_name,
            'row_count': row_count,
            'updated_column': updated_col,
            'max_updated_at': max_updated,
            'created_column': created_col,
            'max_created_at': max_created
        })
    
    return summaries


def create_excel_report(summaries, output_filename):
    """Create Excel report with table summaries"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tables Summary"
    
    # Define headers
    headers = [
        'Table Name',
        'Row Count',
        'Updated Column',
        'Max Updated At',
        'Created Column',
        'Max Created At'
    ]
    
    # Header formatting
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    for row_idx, summary in enumerate(summaries, start=2):
        ws.cell(row=row_idx, column=1, value=summary['table_name'])
        ws.cell(row=row_idx, column=2, value=summary['row_count'])
        ws.cell(row=row_idx, column=3, value=summary['updated_column'])
        ws.cell(row=row_idx, column=4, value=summary['max_updated_at'])
        ws.cell(row=row_idx, column=5, value=summary['created_column'])
        ws.cell(row=row_idx, column=6, value=summary['max_created_at'])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                cell_value = str(cell.value) if cell.value else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_filename)
    print(f"\n✓ Excel report saved: {output_filename}")


def main():
    """Main execution function"""
    print("=" * 70)
    print("MySQL Database Tables Summary Generator")
    print("=" * 70)
    
    # Establish connection
    connection = establish_connection_mysql_db()
    if not connection:
        print("\n✗ Failed to establish database connection. Exiting.")
        return
    
    try:
        # Collect summaries
        summaries = collect_table_summaries(connection)
        
        if not summaries:
            print("\n✗ No tables found or error collecting data.")
            return
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        db_name = DB_CONFIGS['database']
        output_filename = f"mysql_tables_summary_{db_name}_{timestamp}.xlsx"
        
        # Create Excel report
        create_excel_report(summaries, output_filename)
        
        print("\n" + "=" * 70)
        print(f"Summary Statistics:")
        print(f"  Total Tables: {len(summaries)}")
        print(f"  Total Rows: {sum(s['row_count'] for s in summaries):,}")
        print(f"  Tables with Updated Timestamp: {sum(1 for s in summaries if s['updated_column'])}")
        print(f"  Tables with Created Timestamp: {sum(1 for s in summaries if s['created_column'])}")
        print("=" * 70)
        
    except Exception as e:
        print(f"\n✗ Error during execution: {e}")
    
    finally:
        if connection:
            connection.close()
            print("\n✓ Database connection closed.")


if __name__ == "__main__":
    main()