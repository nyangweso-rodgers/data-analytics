import os
import pymysql
from pymysql import Error
from dotenv import load_dotenv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import argparse

# Load environment variables
load_dotenv()


# ============================================================================
# CONFIGURATION
# ============================================================================
DB_CONFIGS = {
    "host": os.getenv("SC_MYSQL_AMTDB_v39_REPLICA_HOST"),
    "port": int(os.getenv("MYSQL_DB_PORT", "3306")),
    "username": os.getenv("SC_MYSQL_AMTDB_v39_REPLICA_USER"),
    "password": os.getenv("SC_MYSQL_AMTDB_v39_REPLICA_PASSWORD", ""),
    "database": os.getenv("SC_MYSQL_AMTDB_v39_REPLICA_DB"),
    "ssl_mode": None,
    "charset": 'utf8mb4'
}


def establish_connection_mysql_db():
    """
    Returns:
        pymysql.connections.Connection: Active database connection or None
    """
    try:
        connection_params = {
            "host": DB_CONFIGS["host"],
            "port": DB_CONFIGS["port"],
            "user": DB_CONFIGS["username"],
            "password": DB_CONFIGS["password"],
            "database": DB_CONFIGS["database"],
            "charset": DB_CONFIGS["charset"],
            "cursorclass": pymysql.cursors.DictCursor
        }
        
        # Only add SSL if ssl_mode is configured
        if DB_CONFIGS["ssl_mode"]:
            connection_params["ssl"] = {'ssl_mode': DB_CONFIGS["ssl_mode"]}
            print("Attempting connection with SSL...")
        else:
            print("Attempting connection without SSL...")
        
        connection = pymysql.connect(**connection_params)
        
        if connection.open:
            ssl_status = "with SSL" if DB_CONFIGS["ssl_mode"] else "without SSL"
            print(f"✓ Successfully connected to database: {DB_CONFIGS['database']} ({ssl_status})")
            return connection
            
    except Error as e:
        print(f"✗ Error connecting to MySQL database: {e}")
        print("\nTroubleshooting tips:")
        print("1. Verify your IP is whitelisted in the database server")
        print("2. Confirm the host IP is correct and accessible")
        print("3. Check username, password, and database name")
        return None


def get_all_tables(connection):
    """Get list of all tables in the database"""
    try:
        with connection.cursor() as cursor:
            cursor.execute("SHOW TABLES")
            tables = [list(row.values())[0] for row in cursor.fetchall()]
        return tables
    except Error as e:
        print(f"✗ Error fetching tables: {e}")
        return []


def get_table_schema(connection, table_name):
    """Get detailed schema information for a table"""
    try:
        with connection.cursor() as cursor:
            # Get column information
            cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
            columns = cursor.fetchall()
            
            schema_info = []
            for col in columns:
                schema_info.append({
                    'column_name': col['Field'],
                    'data_type': col['Type'],
                    'nullable': 'YES' if col['Null'] == 'YES' else 'NO'
                })
            
        return schema_info
    except Error as e:
        print(f"✗ Error getting schema for {table_name}: {e}")
        return []


def get_table_row_count(connection, table_name, use_approximate=False):
    """
    Get row count for a specific table
    
    Args:
        use_approximate: If True, uses information_schema (fast but approximate)
                        If False, uses COUNT(*) (accurate but slow for large tables)
    """
    try:
        with connection.cursor() as cursor:
            if use_approximate:
                query = """
                    SELECT TABLE_ROWS 
                    FROM information_schema.TABLES 
                    WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
                """
                cursor.execute(query, (DB_CONFIGS['database'], table_name))
                result = cursor.fetchone()
                return result['TABLE_ROWS'] if result else 0
            else:
                cursor.execute(f"SELECT COUNT(*) as count FROM `{table_name}`")
                result = cursor.fetchone()
                return result['count']
    except Error as e:
        print(f"✗ Error getting row count for {table_name}: {e}")
        return 0


def find_timestamp_column(columns, patterns):
    """Find first matching timestamp column from list of patterns"""
    for col in columns:
        col_name = col['column_name']
        if col_name.lower() in [p.lower() for p in patterns]:
            return col_name
    return None


def get_max_timestamp(connection, table_name, column_name):
    """Get maximum value from a timestamp column"""
    if not column_name:
        return None
    
    try:
        with connection.cursor() as cursor:
            query = f"SELECT MAX(`{column_name}`) as max_ts FROM `{table_name}`"
            cursor.execute(query)
            result = cursor.fetchone()
            return result['max_ts'] if result['max_ts'] else None
    except Error as e:
        print(f"✗ Error getting max timestamp for {table_name}.{column_name}: {e}")
        return None


def collect_table_details(connection, table_name):
    """Collect detailed information for a single table"""
    print(f"  - Processing: {table_name}")
    
    db_name = DB_CONFIGS['database']
    
    # Get schema info
    schema_info = get_table_schema(connection, table_name)
    
    # Get row count (using approximate for speed)
    row_count = get_table_row_count(connection, table_name, use_approximate=True)
    
    # Find timestamp columns
    updated_patterns = ['updated_at', 'updatedAt', 'modified_at', 'modifiedAt', 'last_updated']
    created_patterns = ['created_at', 'createdAt', 'created', 'date_created']
    
    updated_col = find_timestamp_column(schema_info, updated_patterns)
    created_col = find_timestamp_column(schema_info, created_patterns)
    
    # Get max timestamps
    max_updated = get_max_timestamp(connection, table_name, updated_col)
    max_created = get_max_timestamp(connection, table_name, created_col)
    
    # Build rows for Excel
    rows = []
    for col in schema_info:
        rows.append({
            'database_name': db_name,
            'table_name': table_name,
            'column_name': col['column_name'],
            'data_type': col['data_type'],
            'nullable': col['nullable'],
            'row_count': row_count,
            'max_updated_at': max_updated,
            'max_created_at': max_created
        })
    
    return rows


def create_excel_report(tables_data, output_filename):
    """Create Excel report with summary sheet and one sheet per table"""
    wb = Workbook()
    
    # ========================================================================
    # CREATE SUMMARY SHEET (First sheet)
    # ========================================================================
    summary_ws = wb.active
    summary_ws.title = "📊 Summary"
    
    # Summary headers
    summary_headers = [
        'Table Name',
        'Total Columns',
        'Row Count',
        'Has Updated TS',
        'Max Updated At',
        'Has Created TS',
        'Max Created At',
        'Details Sheet'
    ]
    
    # Header formatting
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Write summary headers
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = summary_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Prepare summary data
    summary_data = []
    for table_name, rows in tables_data.items():
        if rows:  # Should always have rows, but safety check
            first_row = rows[0]
            summary_data.append({
                'table_name': table_name,
                'column_count': len(rows),
                'row_count': first_row['row_count'],
                'has_updated': 'Yes' if first_row['max_updated_at'] else 'No',
                'max_updated_at': first_row['max_updated_at'],
                'has_created': 'Yes' if first_row['max_created_at'] else 'No',
                'max_created_at': first_row['max_created_at'],
                'sheet_name': table_name[:31]  # Excel sheet name limit
            })
    
    # Write summary data
    for row_idx, summary in enumerate(summary_data, start=2):
        summary_ws.cell(row=row_idx, column=1, value=summary['table_name'])
        summary_ws.cell(row=row_idx, column=2, value=summary['column_count'])
        summary_ws.cell(row=row_idx, column=3, value=summary['row_count'])
        summary_ws.cell(row=row_idx, column=4, value=summary['has_updated'])
        summary_ws.cell(row=row_idx, column=5, value=summary['max_updated_at'])
        summary_ws.cell(row=row_idx, column=6, value=summary['has_created'])
        summary_ws.cell(row=row_idx, column=7, value=summary['max_created_at'])
        summary_ws.cell(row=row_idx, column=8, value=f"→ {summary['sheet_name']}")
    
    # Auto-adjust summary sheet column widths
    for column in summary_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                cell_value = str(cell.value) if cell.value else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        summary_ws.column_dimensions[column_letter].width = adjusted_width
    
    # ========================================================================
    # CREATE INDIVIDUAL TABLE SHEETS
    # ========================================================================
    
    # Detail headers
    detail_headers = [
        'Database Name',
        'Table Name',
        'Column Name',
        'Data Type',
        'Nullable',
        'Row Count',
        'Max Updated At',
        'Max Created At'
    ]
    
    # Create a sheet for each table
    for table_name, rows in tables_data.items():
        # Create sheet (Excel sheet names max 31 chars)
        sheet_name = table_name[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        # Write headers
        for col_idx, header in enumerate(detail_headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_idx, row_data in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=row_data['database_name'])
            ws.cell(row=row_idx, column=2, value=row_data['table_name'])
            ws.cell(row=row_idx, column=3, value=row_data['column_name'])
            ws.cell(row=row_idx, column=4, value=row_data['data_type'])
            ws.cell(row=row_idx, column=5, value=row_data['nullable'])
            ws.cell(row=row_idx, column=6, value=row_data['row_count'])
            ws.cell(row=row_idx, column=7, value=row_data['max_updated_at'])
            ws.cell(row=row_idx, column=8, value=row_data['max_created_at'])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_filename)
    print(f"\n✓ Excel report saved: {output_filename}")

def parse_arguments():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(
        description='Generate MySQL database table summaries',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Scan all tables
  python amtdb_tables_summary.py --tables --all
  
  # Scan specific tables
  python amtdb_tables_summary.py --tables users,orders,products
  
  # Scan specific tables (alternative syntax)
  python amtdb_tables_summary.py --tables users orders products
        """
    )
    
    parser.add_argument(
        '--tables',
        type=str,
        nargs='*',
        help='Specify table names (comma-separated or space-separated) or use --all'
    )
    
    parser.add_argument(
        '--all',
        action='store_true',
        help='Scan all tables in the database'
    )
    
    return parser.parse_args()


def parse_table_list(tables_arg):
    """Parse table list from argument (handles comma-separated or space-separated)"""
    if not tables_arg:
        return []
    
    # If single item with commas, split by comma
    if len(tables_arg) == 1 and ',' in tables_arg[0]:
        return [t.strip() for t in tables_arg[0].split(',') if t.strip()]
    
    # Otherwise treat as space-separated list
    return [t.strip() for t in tables_arg if t.strip()]


def main():
    """Main execution function"""
    print("=" * 70)
    print("MySQL Database Tables Summary Generator")
    print("=" * 70)
    
    # Parse arguments
    args = parse_arguments()
    
    # Establish connection
    connection = establish_connection_mysql_db()
    if not connection:
        print("\n✗ Failed to establish database connection. Exiting.")
        return
    
    try:
        # Determine which tables to process
        if args.all:
            tables_to_process = get_all_tables(connection)
            print(f"\n✓ Found {len(tables_to_process)} tables to process (--all)")
        elif args.tables is not None:
            specified_tables = parse_table_list(args.tables)
            if not specified_tables:
                print("\n✗ No tables specified. Use --all or provide table names.")
                return
            
            # Verify tables exist
            all_tables = get_all_tables(connection)
            tables_to_process = []
            for table in specified_tables:
                if table in all_tables:
                    tables_to_process.append(table)
                else:
                    print(f"⚠ Warning: Table '{table}' not found in database")
            
            if not tables_to_process:
                print("\n✗ None of the specified tables exist in the database.")
                return
            
            print(f"\n✓ Processing {len(tables_to_process)} specified table(s)")
        else:
            print("\n✗ Please specify either --all or --tables <table_names>")
            print("Run with --help for usage examples")
            return
        
        # Collect data for each table
        tables_data = {}
        print(f"\nCollecting data from {len(tables_to_process)} table(s)...")
        
        for idx, table_name in enumerate(tables_to_process, 1):
            print(f"[{idx}/{len(tables_to_process)}] ", end="")
            rows = collect_table_details(connection, table_name)
            tables_data[table_name] = rows
        
        if not tables_data:
            print("\n✗ No data collected.")
            return
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        db_name = DB_CONFIGS['database']
        output_filename = f"mysql_tables_summary_{db_name}_{timestamp}.xlsx"
        
        # Create Excel report
        create_excel_report(tables_data, output_filename)
        
        # Summary statistics
        total_columns = sum(len(rows) for rows in tables_data.values())
        
        print("\n" + "=" * 70)
        print(f"Summary Statistics:")
        print(f"  Database: {db_name}")
        print(f"  Tables Processed: {len(tables_data)}")
        print(f"  Total Columns: {total_columns}")
        print(f"  Output File: {output_filename}")
        print("=" * 70)
        
    except Exception as e:
        print(f"\n✗ Error during execution: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        if connection:
            connection.close()
            print("\n✓ Database connection closed.")


if __name__ == "__main__":
    main()