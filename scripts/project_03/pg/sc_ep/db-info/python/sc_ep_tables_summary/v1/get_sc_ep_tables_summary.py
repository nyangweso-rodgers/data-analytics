import os
import psycopg2
from psycopg2 import Error
from dotenv import load_dotenv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import argparse

# Load environment variables
load_dotenv()


# ============================================================================
# CONFIGURATION
# ============================================================================
DB_CONFIGS = {
    "host": os.getenv("SC_EP_PG_DB_HOST"),
    "port": int(os.getenv("PG_DB_PORT", "5432")),
    "username": os.getenv("SC_EP_PG_DB_USER"),
    "password": os.getenv("SC_EP_PG_DB_PASSWORD"),
    "database": os.getenv("SC_EP_PG_DB_NAME"),
    "schema": "public",  # PostgreSQL uses schemas
    "sslmode": "prefer"  # prefer, require, disable
}


def establish_connection_postgres_db():
    """
    Returns:
        psycopg2.connection: Active database connection or None
    """
    try:
        connection_params = {
            "host": DB_CONFIGS["host"],
            "port": DB_CONFIGS["port"],
            "user": DB_CONFIGS["username"],
            "password": DB_CONFIGS["password"],
            "database": DB_CONFIGS["database"],
            "sslmode": DB_CONFIGS["sslmode"]
        }
        
        ssl_status = f"with SSL mode: {DB_CONFIGS['sslmode']}"
        print(f"Attempting connection {ssl_status}...")
        
        connection = psycopg2.connect(**connection_params)
        
        print(f"✓ Successfully connected to database: {DB_CONFIGS['database']}")
        print(f"  Schema: {DB_CONFIGS['schema']}")
        return connection
            
    except Error as e:
        print(f"✗ Error connecting to PostgreSQL database: {e}")
        print("\nTroubleshooting tips:")
        print("1. Verify your IP is whitelisted in pg_hba.conf")
        print("2. Confirm the host IP is correct and accessible")
        print("3. Check username, password, and database name")
        print("4. Verify the schema exists")
        return None


def get_all_tables(connection, schema='public'):
    """Get list of all tables in the specified schema"""
    try:
        cursor = connection.cursor()
        query = """
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = %s 
            AND table_type = 'BASE TABLE'
            ORDER BY table_name
        """
        cursor.execute(query, (schema,))
        tables = [row[0] for row in cursor.fetchall()]
        cursor.close()
        return tables
    except Error as e:
        print(f"✗ Error fetching tables: {e}")
        return []


def get_table_schema(connection, table_name, schema='public'):
    """
    Get detailed schema information for a table
    
    IMPORTANT: Uses udt_name for accurate PostgreSQL type information,
    especially for arrays (_int4, _uuid, etc.) and JSON types
    
    Returns:
        List of dicts with keys: column_name, data_type, udt_name, nullable
    """
    try:
        cursor = connection.cursor()
        query = """
            SELECT 
                column_name,
                data_type,
                udt_name,
                character_maximum_length,
                numeric_precision,
                numeric_scale,
                is_nullable
            FROM information_schema.columns
            WHERE table_schema = %s 
            AND table_name = %s
            ORDER BY ordinal_position
        """
        cursor.execute(query, (schema, table_name))
        columns = cursor.fetchall()
        
        schema_info = []
        for col in columns:
            column_name = col[0]
            data_type = col[1]
            udt_name = col[2]
            char_length = col[3]
            numeric_precision = col[4]
            numeric_scale = col[5]
            is_nullable = col[6]
            
            # ================================================================
            # CRITICAL FIX: Use udt_name for arrays and special types
            # ================================================================
            # PostgreSQL arrays use underscore prefix: _int4, _uuid, _text, etc.
            # JSON types: json, jsonb
            # UUIDs: uuid
            # Custom types: enum names, composite types, etc.
            
            if data_type == 'ARRAY':
                # Use the actual array type (e.g., _int4, _uuid, _text)
                # This is what your YAML schemas need!
                actual_type = udt_name
            elif data_type == 'USER-DEFINED':
                # Custom types (enums, composite types, etc.)
                actual_type = udt_name
            elif udt_name in ('json', 'jsonb', 'uuid'):
                # Use udt_name for JSON and UUID types
                actual_type = udt_name
            else:
                # For standard types (varchar, integer, etc.), use data_type
                actual_type = data_type
            
            # Build complete data type string with parameters
            if char_length:
                # varchar(255), char(10), etc.
                actual_type = f"{actual_type}({char_length})"
            elif numeric_precision:
                if numeric_scale:
                    # decimal(18,2), numeric(10,4), etc.
                    actual_type = f"{actual_type}({numeric_precision},{numeric_scale})"
                else:
                    # decimal(18), numeric(10), etc.
                    actual_type = f"{actual_type}({numeric_precision})"
            
            schema_info.append({
                'column_name': column_name,
                'data_type': actual_type,  # This now correctly shows _int4 instead of ARRAY
                'udt_name': udt_name,  # Keep raw udt_name for reference
                'nullable': is_nullable
            })
        
        cursor.close()
        return schema_info
    except Error as e:
        print(f"✗ Error getting schema for {table_name}: {e}")
        return []


def get_table_row_count(connection, table_name, schema='public', use_approximate=False):
    """
    Get row count for a specific table
    
    Args:
        use_approximate: If True, uses pg_class statistics (fast but approximate)
                        If False, uses COUNT(*) (accurate but slow for large tables)
    """
    try:
        cursor = connection.cursor()
        if use_approximate:
            # Use PostgreSQL statistics (much faster, but may be outdated)
            query = """
                SELECT n_live_tup 
                FROM pg_stat_user_tables 
                WHERE schemaname = %s AND relname = %s
            """
            cursor.execute(query, (schema, table_name))
            result = cursor.fetchone()
            count = result[0] if result else 0
        else:
            # Exact count (slower)
            query = f'SELECT COUNT(*) FROM "{schema}"."{table_name}"'
            cursor.execute(query)
            result = cursor.fetchone()
            count = result[0] if result else 0
        
        cursor.close()
        return count
    except Error as e:
        print(f"✗ Error getting row count for {table_name}: {e}")
        return 0


def find_timestamp_column(columns, patterns):
    """Find first matching timestamp column from list of patterns"""
    for col in columns:
        col_name = col['column_name']
        if col_name.lower() in [p.lower() for p in patterns]:
            return col_name
    return None


def get_max_timestamp(connection, table_name, column_name, schema='public'):
    """Get maximum value from a timestamp column"""
    if not column_name:
        return None
    
    try:
        cursor = connection.cursor()
        query = f'SELECT MAX("{column_name}") FROM "{schema}"."{table_name}"'
        cursor.execute(query)
        result = cursor.fetchone()
        cursor.close()
        return result[0] if result[0] else None
    except Error as e:
        print(f"✗ Error getting max timestamp for {table_name}.{column_name}: {e}")
        return None


def collect_table_details(connection, table_name, schema='public'):
    """Collect detailed information for a single table"""
    print(f"  - Processing: {schema}.{table_name}")
    
    db_name = DB_CONFIGS['database']
    
    # Get schema info
    schema_info = get_table_schema(connection, table_name, schema)
    
    # Get row count (using approximate for speed)
    row_count = get_table_row_count(connection, table_name, schema, use_approximate=True)
    
    # Find timestamp columns
    updated_patterns = ['updated_at', 'updatedAt', 'modified_at', 'modifiedAt', 'last_updated']
    created_patterns = ['created_at', 'createdAt', 'created', 'date_created']
    
    updated_col = find_timestamp_column(schema_info, updated_patterns)
    created_col = find_timestamp_column(schema_info, created_patterns)
    
    # Get max timestamps
    max_updated = get_max_timestamp(connection, table_name, updated_col, schema)
    max_created = get_max_timestamp(connection, table_name, created_col, schema)
    
    # Build rows for Excel
    rows = []
    for col in schema_info:
        rows.append({
            'database_name': db_name,
            'schema_name': schema,
            'table_name': table_name,
            'column_name': col['column_name'],
            'data_type': col['data_type'],  # Now correctly shows _int4, _uuid, etc.
            'udt_name': col['udt_name'],  # Raw PostgreSQL type name
            'nullable': col['nullable'],
            'row_count': row_count,
            'max_updated_at': max_updated,
            'max_created_at': max_created
        })
    
    return rows


def create_excel_report(tables_data, output_filename, schema='public'):
    """Create Excel report with summary sheet and one sheet per table"""
    wb = Workbook()
    
    # ========================================================================
    # CREATE SUMMARY SHEET (First sheet)
    # ========================================================================
    summary_ws = wb.active
    summary_ws.title = "📊 Summary"
    
    # Summary headers
    summary_headers = [
        'Schema',
        'Table Name',
        'Total Columns',
        'Row Count',
        'Has Updated TS',
        'Max Updated At',
        'Has Created TS',
        'Max Created At',
        'Details Sheet'
    ]
    
    # Header formatting
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Write summary headers
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = summary_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Prepare summary data
    summary_data = []
    for table_name, rows in tables_data.items():
        if rows:  # Should always have rows, but safety check
            first_row = rows[0]
            summary_data.append({
                'schema': first_row['schema_name'],
                'table_name': table_name,
                'column_count': len(rows),
                'row_count': first_row['row_count'],
                'has_updated': 'Yes' if first_row['max_updated_at'] else 'No',
                'max_updated_at': first_row['max_updated_at'],
                'has_created': 'Yes' if first_row['max_created_at'] else 'No',
                'max_created_at': first_row['max_created_at'],
                'sheet_name': table_name[:31]  # Excel sheet name limit
            })
    
    # Write summary data
    for row_idx, summary in enumerate(summary_data, start=2):
        summary_ws.cell(row=row_idx, column=1, value=summary['schema'])
        summary_ws.cell(row=row_idx, column=2, value=summary['table_name'])
        summary_ws.cell(row=row_idx, column=3, value=summary['column_count'])
        summary_ws.cell(row=row_idx, column=4, value=summary['row_count'])
        summary_ws.cell(row=row_idx, column=5, value=summary['has_updated'])
        summary_ws.cell(row=row_idx, column=6, value=summary['max_updated_at'])
        summary_ws.cell(row=row_idx, column=7, value=summary['has_created'])
        summary_ws.cell(row=row_idx, column=8, value=summary['max_created_at'])
        summary_ws.cell(row=row_idx, column=9, value=f"→ {summary['sheet_name']}")
    
    # Add totals row to summary sheet
    if summary_data:
        totals_row = len(summary_data) + 2
        summary_ws.cell(row=totals_row, column=2, value="TOTALS").font = Font(bold=True)
        summary_ws.cell(row=totals_row, column=3, value=sum(s['column_count'] for s in summary_data)).font = Font(bold=True)
        summary_ws.cell(row=totals_row, column=4, value=sum(s['row_count'] for s in summary_data)).font = Font(bold=True)
        summary_ws.cell(row=totals_row, column=5, value=f"{sum(1 for s in summary_data if s['has_updated'] == 'Yes')}/{len(summary_data)}").font = Font(bold=True)
        summary_ws.cell(row=totals_row, column=7, value=f"{sum(1 for s in summary_data if s['has_created'] == 'Yes')}/{len(summary_data)}").font = Font(bold=True)
    
    # Auto-adjust summary sheet column widths
    for column in summary_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                cell_value = str(cell.value) if cell.value else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        summary_ws.column_dimensions[column_letter].width = adjusted_width
    
    # ========================================================================
    # CREATE INDIVIDUAL TABLE SHEETS
    # ========================================================================
    
    # Detail headers - NOW INCLUDES UDT_NAME
    detail_headers = [
        'Database Name',
        'Schema Name',
        'Table Name',
        'Column Name',
        'Data Type',
        'UDT Name',  # Added this for debugging
        'Nullable',
        'Row Count',
        'Max Updated At',
        'Max Created At'
    ]
    
    # Create a sheet for each table
    for table_name, rows in tables_data.items():
        # Create sheet (Excel sheet names max 31 chars)
        sheet_name = table_name[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        # Write headers
        for col_idx, header in enumerate(detail_headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_idx, row_data in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=row_data['database_name'])
            ws.cell(row=row_idx, column=2, value=row_data['schema_name'])
            ws.cell(row=row_idx, column=3, value=row_data['table_name'])
            ws.cell(row=row_idx, column=4, value=row_data['column_name'])
            ws.cell(row=row_idx, column=5, value=row_data['data_type'])
            ws.cell(row=row_idx, column=6, value=row_data['udt_name'])  # Raw PostgreSQL type
            ws.cell(row=row_idx, column=7, value=row_data['nullable'])
            ws.cell(row=row_idx, column=8, value=row_data['row_count'])
            ws.cell(row=row_idx, column=9, value=row_data['max_updated_at'])
            ws.cell(row=row_idx, column=10, value=row_data['max_created_at'])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_filename)
    print(f"\n✓ Excel report saved: {output_filename}")


def parse_arguments():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(
        description='Generate PostgreSQL database table summaries',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Scan all tables in default schema (public)
  python postgres_tables_summary.py --tables --all
  
  # Scan all tables in specific schema
  python postgres_tables_summary.py --tables --all --schema myschema
  
  # Scan specific tables
  python postgres_tables_summary.py --tables users,orders,products
  
  # Scan specific tables (alternative syntax)
  python postgres_tables_summary.py --tables users orders products

IMPORTANT NOTE:
  This script now correctly reports PostgreSQL array types using udt_name.
  For example, an integer array will show as "_int4" instead of "ARRAY".
  Use this "Data Type" column value directly in your YAML schemas!
        """
    )
    
    parser.add_argument(
        '--tables',
        type=str,
        nargs='*',
        help='Specify table names (comma-separated or space-separated) or use --all'
    )
    
    parser.add_argument(
        '--all',
        action='store_true',
        help='Scan all tables in the database'
    )
    
    parser.add_argument(
        '--schema',
        type=str,
        default=None,
        help='PostgreSQL schema name (default: from .env or "public")'
    )
    
    return parser.parse_args()


def parse_table_list(tables_arg):
    """Parse table list from argument (handles comma-separated or space-separated)"""
    if not tables_arg:
        return []
    
    # If single item with commas, split by comma
    if len(tables_arg) == 1 and ',' in tables_arg[0]:
        return [t.strip() for t in tables_arg[0].split(',') if t.strip()]
    
    # Otherwise treat as space-separated list
    return [t.strip() for t in tables_arg if t.strip()]


def main():
    """Main execution function"""
    print("=" * 70)
    print("PostgreSQL Database Tables Summary Generator")
    print("=" * 70)
    
    # Parse arguments
    args = parse_arguments()
    
    # Determine schema to use
    schema = args.schema if args.schema else DB_CONFIGS['schema']
    
    # Establish connection
    connection = establish_connection_postgres_db()
    if not connection:
        print("\n✗ Failed to establish database connection. Exiting.")
        return
    
    try:
        # Determine which tables to process
        if args.all:
            tables_to_process = get_all_tables(connection, schema)
            print(f"\n✓ Found {len(tables_to_process)} tables in schema '{schema}' (--all)")
        elif args.tables is not None:
            specified_tables = parse_table_list(args.tables)
            if not specified_tables:
                print("\n✗ No tables specified. Use --all or provide table names.")
                return
            
            # Verify tables exist
            all_tables = get_all_tables(connection, schema)
            tables_to_process = []
            for table in specified_tables:
                if table in all_tables:
                    tables_to_process.append(table)
                else:
                    print(f"⚠ Warning: Table '{table}' not found in schema '{schema}'")
            
            if not tables_to_process:
                print(f"\n✗ None of the specified tables exist in schema '{schema}'.")
                return
            
            print(f"\n✓ Processing {len(tables_to_process)} specified table(s) from schema '{schema}'")
        else:
            print("\n✗ Please specify either --all or --tables <table_names>")
            print("Run with --help for usage examples")
            return
        
        # Collect data for each table
        tables_data = {}
        print(f"\nCollecting data from {len(tables_to_process)} table(s)...")
        
        for idx, table_name in enumerate(tables_to_process, 1):
            print(f"[{idx}/{len(tables_to_process)}] ", end="")
            rows = collect_table_details(connection, table_name, schema)
            tables_data[table_name] = rows
        
        if not tables_data:
            print("\n✗ No data collected.")
            return
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        db_name = DB_CONFIGS['database']
        output_filename = f"postgres_tables_summary_{db_name}_{schema}_{timestamp}.xlsx"
        
        # Create Excel report
        create_excel_report(tables_data, output_filename, schema)
        
        # Summary statistics
        total_columns = sum(len(rows) for rows in tables_data.values())
        
        print("\n" + "=" * 70)
        print(f"Summary Statistics:")
        print(f"  Database: {db_name}")
        print(f"  Schema: {schema}")
        print(f"  Tables Processed: {len(tables_data)}")
        print(f"  Total Columns: {total_columns}")
        print(f"  Output File: {output_filename}")
        print("=" * 70)
        print("\n✅ IMPORTANT: The 'Data Type' column now shows accurate types!")
        print("   - Arrays: _int4, _uuid, _text (use these in your YAML)")
        print("   - JSON: json, jsonb (use these in your YAML)")
        print("   - UUIDs: uuid (use this in your YAML)")
        print("=" * 70)
        
    except Exception as e:
        print(f"\n✗ Error during execution: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        if connection:
            connection.close()
            print("\n✓ Database connection closed.")


if __name__ == "__main__":
    main()